"""Recalculate sprint-queue.xlsx Dashboard KPIs from the Catalog, in pure Python.

Matches the v2 workbook structure (4 sheets; 19-column Catalog table `Table1`;
Variables + Sprint Details sheets). Reads the Catalog by HEADER name (so column
order is irrelevant), then writes literal values into the documented Dashboard
cells so ``data_only`` reads are correct without opening Excel. Usage:

    python recalc.py <path-to-xlsx>

Note: this overwrites the Dashboard's live formulas with computed literals. It is
for headless refresh; in normal Excel use the formulas recompute on their own.
"""
import sys
from openpyxl import load_workbook

# LOE points (the Dashboard "Effort by LOE" helper table, D12:E19)
LOE_POINTS = {"XS": 0.5, "XS-S": 0.75, "S": 1, "S-M": 2, "M": 3, "M-L": 5, "L": 8, "XL": 20}
# Dashboard row for each LOE size in the helper table (col D=size, F/G/H computed)
SIZE_ROW = {"XS": 12, "XS-S": 13, "S": 14, "S-M": 15, "M": 16, "M-L": 17, "L": 18, "XL": 19}
DONE_EXACT = {"Dissolved", "Superseded"}


def _rows(ws):
    headers = [c.value for c in ws[1]]
    idx = {(h.strip() if isinstance(h, str) else h): i for i, h in enumerate(headers) if h}

    def cell(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = cell(row, "Sprint Code")
        if code in (None, ""):
            continue

        def s(name):
            v = cell(row, name)
            return v.strip() if isinstance(v, str) else ("" if v is None else v)

        out.append({"code": code, "phase": s("Phase"), "loe": s("LOE"),
                    "impl": s("Implementation Status"), "written": s("Written Status")})
    return out


def _is_completed(impl):
    return isinstance(impl, str) and impl.startswith("Completed")


def _is_done(impl):
    return _is_completed(impl) or impl in DONE_EXACT


def _find_catalog(wb):
    """Locate the sprint catalog sheet by name, else by a 'Sprint Code' header
    (robust to the data sheet being named Catalog, DATA.sprint-catalog, etc.)."""
    for name in wb.sheetnames:
        if name.strip().lower() in ("catalog", "data.sprint-catalog", "sprint catalog"):
            return wb[name]
    for ws in wb.worksheets:
        if any(isinstance(c.value, str) and c.value.strip() == "Sprint Code" for c in ws[1]):
            return ws
    raise KeyError("No catalog sheet (need a 'Sprint Code' header in row 1)")


def recalc(path):
    wb = load_workbook(path)
    cat = _find_catalog(wb)
    dash = wb["Dashboard"]
    rows = _rows(cat)

    total = len(rows)
    completed = sum(1 for r in rows if _is_completed(r["impl"]))
    blocked = sum(1 for r in rows if r["impl"] == "Blocked")
    queued = sum(1 for r in rows if r["impl"] == "Queued")
    in_flight = sum(1 for r in rows if r["impl"] == "In Flight")
    dissolved = sum(1 for r in rows if r["impl"] == "Dissolved")
    superseded = sum(1 for r in rows if r["impl"] == "Superseded")
    eff_total = total - dissolved - superseded
    pct_count = (completed / eff_total) if eff_total else 0

    # Effort-by-LOE helper table (writes F/G/H per size + totals on row 20)
    rem_pts = comp_pts = 0.0
    rem_cnt = 0
    for size, pts in LOE_POINTS.items():
        not_done = sum(1 for r in rows if r["loe"] == size and not _is_done(r["impl"]))
        comp = sum(1 for r in rows if r["loe"] == size and _is_completed(r["impl"]))
        rr = SIZE_ROW[size]
        dash.cell(rr, 6).value = not_done          # F: Remaining #
        dash.cell(rr, 7).value = not_done * pts     # G: Rem Pts
        dash.cell(rr, 8).value = comp * pts         # H: Comp Pts
        rem_cnt += not_done
        rem_pts += not_done * pts
        comp_pts += comp * pts
    dash["F20"] = rem_cnt
    dash["G20"] = rem_pts
    dash["H20"] = comp_pts

    total_loe = rem_pts + comp_pts
    pct_loe = (comp_pts / total_loe) if total_loe else 0
    sprints_remaining = blocked + queued + in_flight
    avg = (total_loe / eff_total) if eff_total else 0

    vals = {
        "B11": total, "B12": completed, "B13": blocked, "B14": queued, "B15": in_flight,
        "B16": dissolved, "B17": superseded, "B18": pct_count,
        "B21": rem_pts, "B22": comp_pts, "B23": total_loe, "B24": pct_loe,
        "B25": sprints_remaining, "B26": avg,
        # Status Distribution helper (drives the doughnut chart)
        "E23": completed, "E24": blocked, "E25": queued, "E26": dissolved, "E27": superseded,
    }
    for cellref, v in vals.items():
        dash[cellref] = v
    wb.save(path)
    return vals


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: python recalc.py <path-to-xlsx>", file=sys.stderr)
        sys.exit(2)
    out = recalc(sys.argv[1])
    print("recalc OK:", {k: out[k] for k in ("B11", "B12", "B23", "B24")})
