import os, subprocess, sys
from openpyxl import load_workbook, Workbook

HERE = os.path.dirname(os.path.abspath(__file__))

def _make_catalog(path):
    wb = Workbook()
    ws = wb.active; ws.title = "Dashboard"
    for cell in ["B11","B12","B13","B14","B15","B16","B17","B20","B21","B22",
                 "B23","B24","B25","B29","B32","B35","B36","B37","B38"]:
        ws[cell] = None
    cat = wb.create_sheet("Catalog")
    cat.append(["Seq","Sprint Code","Title","Phase","Stage","LOE","Deps",
                "Implementation Status","Written Status"])
    rows = [
        [1, "SK-01", "A", "Phase 1", "",  "M",  "", "Queued",    "Full Spec"],
        [2, "SK-02", "B", "Phase 1", "",  "S",  "", "Queued",    "Stub"],
        [3, "SK-03", "C", "Phase 1", "",  "L",  "", "Completed", "Full Spec"],
        ["","SK-04", "D", "Phase 2", "",  "XL", "", "Dissolved", "None"],
    ]
    for r in rows:
        cat.append(r)
    wb.save(path)

def test_recalc_populates_dashboard(tmp_path):
    p = str(tmp_path / "q.xlsx")
    _make_catalog(p)
    subprocess.run([sys.executable, os.path.join(HERE, "recalc.py"), p], check=True)
    wb = load_workbook(p)
    d = wb["Dashboard"]
    assert d["B11"].value == 3            # total excl. dissolved
    assert d["B12"].value == 1            # completed
    assert d["B14"].value == 2            # queued
    assert d["B23"].value == 12           # total LOE excl dissolved: M3 + S1 + L8
    assert d["B22"].value == 8            # completed LOE: L8
    assert round(d["B24"].value, 4) == round(8 / 12, 4)  # % complete by LOE
    assert d["B29"].value == 1            # full specs queued (SK-01 only)
    assert d["B35"].value == 3            # sprints in current phase (Phase 1)
