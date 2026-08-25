#!/usr/bin/env python3
"""Generate the human-facing spreadsheet report from the live work register.

This is a **presentation output** (items 58, 84). It is written by its registered
generator and never read back as operational truth: every figure it shows is a
computed literal stamped with the provider, source, and snapshot time it came
from, so a stale cache cannot silently contradict the register.

The target role must be registered with ``origin: generated`` (or
``mutability: generated``); the script refuses to write over an authored or
authoritative document.

Usage:
    build_register_report.py --root <dir> [--role sprintQueue] [--out <path>]
"""
from __future__ import annotations

import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tools.governance import policy as policy_mod, providers, registry as registry_mod  # noqa: E402
from tools.governance.errors import GovernanceError, RoleNotRegistered  # noqa: E402
from tools.governance.providers import kpi  # noqa: E402

DEFAULT_ROLE = "sprintQueue"
COLUMNS = [
    ("sequence", "Seq"), ("id", "Id"), ("title", "Title"), ("group", "Group"),
    ("lane", "Lane"), ("effort", "Effort"), ("prerequisites", "Prerequisites"),
    ("raw_status", "Status"), ("status", "Status (canonical)"),
    ("raw_written_status", "Specification"), ("branch", "Branch"),
    ("started", "Started"), ("completed", "Completed"), ("evidence", "Evidence"),
    ("spec_link", "Specification link"), ("description", "Description"), ("notes", "Notes"),
]


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise GovernanceError(
            "generating the spreadsheet report needs openpyxl, which is not installed (%s). "
            "Install it, or drop the generated report role from the registry." % exc,
            detail={"dependency": "openpyxl"}) from exc
    return openpyxl


def resolve_target(reg, role_name: str, override: str) -> str:
    if override:
        return override if os.path.isabs(override) else os.path.join(reg.root, override)
    spec = reg.roles.get(role_name)
    if spec is None:
        raise RoleNotRegistered(
            "role %r is not registered, so there is no report to generate. Register it "
            "(provider xlsx, authority report, mutability generated, generatedFrom "
            "workRegister) or pass --out." % role_name, detail={"role": role_name})
    if spec.is_external:
        raise GovernanceError("role %r is external; this generator writes local files only"
                              % role_name)
    if spec.origin != "generated" and spec.mutability != "generated":
        raise GovernanceError(
            "role %r is registered as origin=%s mutability=%s. This generator only writes "
            "roles the project has declared generated — it will not overwrite an authored "
            "or authoritative document." % (role_name, spec.origin, spec.mutability),
            detail={"role": role_name})
    return os.path.join(reg.root, *spec.path.split("/"))


def build(root: str, *, role_name: str = DEFAULT_ROLE, out: str = "") -> str:
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    reg = registry_mod.load(root)
    project_policy = policy_mod.load(reg.policy)
    selection = providers.work_register(reg, actor="roadmap-review")
    snapshot = selection.provider.snapshot()
    metrics = kpi.compute(snapshot, effort_scale=project_policy.get("roadmap.effortScale"),
                          dispatch_buffer=project_policy.dispatch_buffer)
    target = resolve_target(reg, role_name, out)

    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E78")

    summary["A1"] = "Work register report"
    summary["A1"].font = Font(bold=True, size=16)
    summary["A1"].alignment = Alignment(vertical="center")

    provenance_rows = [
        ("Generated from", snapshot.source),
        ("Provider", snapshot.provider),
        ("Registered role", "%s (%s)" % (selection.role_name, selection.authority)),
        ("Snapshot taken", snapshot.taken_at),
        ("Snapshot stale", "YES — %s" % snapshot.stale_reason if snapshot.stale else "no"),
        ("Read via compatibility adapter", "yes" if selection.compatibility else "no"),
    ]
    row = 3
    for label, value in provenance_rows:
        summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        summary.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    summary.cell(row=row, column=1, value="Metric").font = header_font
    summary.cell(row=row, column=1).fill = header_fill
    summary.cell(row=row, column=2, value="Value").font = header_font
    summary.cell(row=row, column=2).fill = header_fill
    row += 1
    for metric in metrics.metrics:
        summary.cell(row=row, column=1, value=metric.name)
        if metric.computable:
            summary.cell(row=row, column=2,
                         value="%s%s" % (metric.value, (" " + metric.unit) if metric.unit else ""))
        else:
            summary.cell(row=row, column=2,
                         value="not computable — missing: %s" % ", ".join(metric.missing_inputs))
        row += 1
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 72

    data = workbook.create_sheet("Items")
    for column, (_field, label) in enumerate(COLUMNS, start=1):
        cell = data.cell(row=1, column=column, value=label)
        cell.font = header_font
        cell.fill = header_fill
    for index, item in enumerate(snapshot.items, start=2):
        payload = item.as_dict()
        for column, (field_name, _label) in enumerate(COLUMNS, start=1):
            value = payload.get(field_name, "")
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            data.cell(row=index, column=column, value=value)
    for column in range(1, len(COLUMNS) + 1):
        data.column_dimensions[get_column_letter(column)].width = 18
    data.freeze_panes = "A2"

    os.makedirs(os.path.dirname(target) or ".", exist_ok=True)
    workbook.save(target)
    return target


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    parser.add_argument("--root", default=os.getcwd())
    parser.add_argument("--role", default=DEFAULT_ROLE)
    parser.add_argument("--out", default="")
    args = parser.parse_args(argv)
    try:
        target = build(args.root, role_name=args.role, out=args.out)
    except GovernanceError as exc:
        print(str(exc), file=sys.stderr)
        return 3
    print("register report written: %s" % target)
    return 0


if __name__ == "__main__":
    sys.exit(main())
