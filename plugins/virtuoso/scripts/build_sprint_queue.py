"""Build sprint-queue.xlsx template with Dashboard + Catalog tabs.

Usage: build_sprint_queue.py [catalog_csv] [output_xlsx]

Writes the workbook to output_xlsx (resolved against the current working
directory), or next to this script when no output path is given. catalog_csv
is accepted for CLI compatibility but not yet read — the Catalog tab is
populated with example rows.
"""
import argparse
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

parser = argparse.ArgumentParser(
    description="Build the sprint-queue.xlsx template (Dashboard + Catalog tabs).")
parser.add_argument(
    "catalog", nargs="?",
    help="sprint catalog CSV (reserved; not yet read — example rows are written)")
parser.add_argument(
    "output", nargs="?",
    help="output xlsx path (default: sprint-queue.xlsx next to this script)")
args = parser.parse_args()

OUT = os.path.abspath(args.output) if args.output else os.path.join(
    SCRIPT_DIR, "sprint-queue.xlsx")

# === Styles ===
FONT_NAME = "Arial"
title_font = Font(name=FONT_NAME, size=18, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", start_color="1F4E78")
section_font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
section_fill = PatternFill("solid", start_color="2E75B6")
label_font = Font(name=FONT_NAME, size=10, bold=True)
value_font = Font(name=FONT_NAME, size=10)
header_font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="1F4E78")
input_font = Font(name=FONT_NAME, size=10, color="0000FF")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# LOE point map: S=1, M=3, L=8, XL=20
LOE_TERM = (
    '((Catalog!F:F="S")*1 + (Catalog!F:F="M")*3 + '
    '(Catalog!F:F="L")*8 + (Catalog!F:F="XL")*20)'
)

wb = Workbook()

# ========== Tab 1: Dashboard ==========
ws = wb.active
ws.title = "Dashboard"

ws["A1"] = "Project Roadmap Dashboard"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws["A1"].alignment = center
ws.merge_cells("A1:F1")
ws.row_dimensions[1].height = 32

# Project Info (rows 3-8)
ws["A3"] = "Project Info"
ws["A3"].font = section_font
ws["A3"].fill = section_fill
ws["A3"].alignment = left
ws.merge_cells("A3:F3")

info_rows = [
    ("Project Name", "[Enter project name]"),
    ("Roadmap Document", "[e.g., GoG_Roadmap.md]"),
    ("Finish Line Target", "[Enter finish line description]"),
    ("Current Phase", "[e.g., Phase 1]"),
    ("Last Roadmap Review", "[YYYY-MM-DD]"),
]
for i, (label, val) in enumerate(info_rows, start=4):
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = label_font
    ws[f"B{i}"] = val
    ws[f"B{i}"].font = input_font

# Pipeline Status (rows 10-17)
PIPELINE_HEADER = 10
ws[f"A{PIPELINE_HEADER}"] = "Pipeline Status"
ws[f"A{PIPELINE_HEADER}"].font = section_font
ws[f"A{PIPELINE_HEADER}"].fill = section_fill
ws.merge_cells(f"A{PIPELINE_HEADER}:F{PIPELINE_HEADER}")

pipeline_rows = [
    ("Total sprints", '=COUNTA(Catalog!B:B)-1'),
    ("Completed", '=COUNTIF(Catalog!H:H,"Completed")'),
    ("In Flight", '=COUNTIF(Catalog!H:H,"In Flight")'),
    ("Queued", '=COUNTIF(Catalog!H:H,"Queued")'),
    ("Blocked", '=COUNTIF(Catalog!H:H,"Blocked")'),
    ("Dissolved", '=COUNTIF(Catalog!H:H,"Dissolved")'),
    ("% Complete (by count)", '=IFERROR(B12/B11,0)'),
]
for i, (label, formula) in enumerate(pipeline_rows, start=PIPELINE_HEADER + 1):
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = label_font
    ws[f"B{i}"] = formula
    ws[f"B{i}"].font = value_font
ws["B17"].number_format = "0.0%"

# Progress to Finish Line (rows 19-25)
PROGRESS_HEADER = 19
ws[f"A{PROGRESS_HEADER}"] = "Progress to Finish Line"
ws[f"A{PROGRESS_HEADER}"].font = section_font
ws[f"A{PROGRESS_HEADER}"].fill = section_fill
ws.merge_cells(f"A{PROGRESS_HEADER}:F{PROGRESS_HEADER}")

loe_remaining = (
    f'=SUMPRODUCT({LOE_TERM} * '
    f'((Catalog!H:H="Queued") + (Catalog!H:H="In Flight") + '
    f'(Catalog!H:H="Blocked")))'
)
loe_completed = f'=SUMPRODUCT({LOE_TERM} * (Catalog!H:H="Completed"))'

progress_rows = [
    ("Sprints remaining (uncompleted, undissolved)", '=B11-B12-B16'),
    ("LOE remaining (points)", loe_remaining),
    ("LOE completed (points)", loe_completed),
    ("Total LOE (points, excl. dissolved)", '=B21+B22'),
    ("% Complete (by LOE)", '=IFERROR(B22/B23,0)'),
    ("Avg sprint size (LOE points)", '=IFERROR(B23/(B11-B16),0)'),
]
for i, (label, formula) in enumerate(progress_rows, start=PROGRESS_HEADER + 1):
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = label_font
    ws[f"B{i}"] = formula
    ws[f"B{i}"].font = value_font
ws["B24"].number_format = "0.0%"

# Eager-Spec Buffer (rows 27-32)
BUFFER_HEADER = 27
ws[f"A{BUFFER_HEADER}"] = "Eager-Spec Buffer Status"
ws[f"A{BUFFER_HEADER}"].font = section_font
ws[f"A{BUFFER_HEADER}"].fill = section_fill
ws.merge_cells(f"A{BUFFER_HEADER}:F{BUFFER_HEADER}")

buffer_rows = [
    ("Target", 5),
    ("Full specs queued",
        '=COUNTIFS(Catalog!H:H,"Queued",Catalog!I:I,"Full Spec")'),
    ("Stubs queued",
        '=COUNTIFS(Catalog!H:H,"Queued",Catalog!I:I,"Stub")'),
    ("Drafts queued",
        '=COUNTIFS(Catalog!H:H,"Queued",Catalog!I:I,"Draft")'),
    ("Buffer health",
        '=IF(B29>=5,"Healthy",IF(B29>=3,"Running low",IF(B29>=1,"Critical","Empty")))'),
]
for i, (label, val) in enumerate(buffer_rows, start=BUFFER_HEADER + 1):
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = label_font
    ws[f"B{i}"] = val
    ws[f"B{i}"].font = value_font

# Current Phase Progress (rows 34-38)
PHASE_HEADER = 34
ws[f"A{PHASE_HEADER}"] = "Current Phase Progress"
ws[f"A{PHASE_HEADER}"].font = section_font
ws[f"A{PHASE_HEADER}"].fill = section_fill
ws.merge_cells(f"A{PHASE_HEADER}:F{PHASE_HEADER}")

phase_rows = [
    ("Sprints in current phase", '=COUNTIF(Catalog!D:D,B7)'),
    ("Completed in current phase",
        '=COUNTIFS(Catalog!D:D,B7,Catalog!H:H,"Completed")'),
    ("Remaining in current phase", '=B35-B36'),
    ("% phase complete", '=IFERROR(B36/B35,0)'),
]
for i, (label, formula) in enumerate(phase_rows, start=PHASE_HEADER + 1):
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = label_font
    ws[f"B{i}"] = formula
    ws[f"B{i}"].font = value_font
ws["B38"].number_format = "0.0%"

# Chart data
CHART_DATA_START = 42
ws[f"A{CHART_DATA_START}"] = "Chart Data (auto-calculated)"
ws[f"A{CHART_DATA_START}"].font = Font(name=FONT_NAME, size=9, italic=True, color="808080")
ws[f"A{CHART_DATA_START + 1}"] = "Status"
ws[f"B{CHART_DATA_START + 1}"] = "Count"
ws[f"A{CHART_DATA_START + 1}"].font = label_font
ws[f"B{CHART_DATA_START + 1}"].font = label_font

status_chart_data = [
    ("Queued", '=COUNTIF(Catalog!H:H,"Queued")'),
    ("In Flight", '=COUNTIF(Catalog!H:H,"In Flight")'),
    ("Blocked", '=COUNTIF(Catalog!H:H,"Blocked")'),
    ("Completed", '=COUNTIF(Catalog!H:H,"Completed")'),
    ("Dissolved", '=COUNTIF(Catalog!H:H,"Dissolved")'),
]
for i, (label, formula) in enumerate(status_chart_data, start=CHART_DATA_START + 2):
    ws[f"A{i}"] = label
    ws[f"B{i}"] = formula

WS_CHART_START = 50
ws[f"A{WS_CHART_START}"] = "Written"
ws[f"B{WS_CHART_START}"] = "Count"
ws[f"A{WS_CHART_START}"].font = label_font
ws[f"B{WS_CHART_START}"].font = label_font

written_chart_data = [
    ("None", '=COUNTIF(Catalog!I:I,"None")'),
    ("Stub", '=COUNTIF(Catalog!I:I,"Stub")'),
    ("Draft", '=COUNTIF(Catalog!I:I,"Draft")'),
    ("Skeleton", '=COUNTIF(Catalog!I:I,"Skeleton")'),
    ("Full Spec", '=COUNTIF(Catalog!I:I,"Full Spec")'),
]
for i, (label, formula) in enumerate(written_chart_data, start=WS_CHART_START + 1):
    ws[f"A{i}"] = label
    ws[f"B{i}"] = formula

# Pie chart
pie = PieChart()
pie.title = "Implementation Status"
labels = Reference(ws, min_col=1, min_row=CHART_DATA_START + 2, max_row=CHART_DATA_START + 6)
data = Reference(ws, min_col=2, min_row=CHART_DATA_START + 1, max_row=CHART_DATA_START + 6)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.dataLabels = DataLabelList(showPercent=True)
pie.height = 9
pie.width = 13
ws.add_chart(pie, "D10")

# Bar chart
bar = BarChart()
bar.type = "bar"
bar.title = "Written Status"
ws_labels = Reference(ws, min_col=1, min_row=WS_CHART_START + 1, max_row=WS_CHART_START + 5)
ws_data = Reference(ws, min_col=2, min_row=WS_CHART_START, max_row=WS_CHART_START + 5)
bar.add_data(ws_data, titles_from_data=True)
bar.set_categories(ws_labels)
bar.height = 9
bar.width = 13
ws.add_chart(bar, "D27")

for col, w in enumerate([34, 22, 4, 18, 18, 18, 18], 1):
    ws.column_dimensions[get_column_letter(col)].width = w

# ========== Tab 2: Catalog ==========
cat = wb.create_sheet("Catalog")

headers = [
    "Seq", "Sprint Code", "Title", "Phase", "Stage", "LOE",
    "Dependencies", "Implementation Status", "Written Status",
    "Description", "Branch", "Date Started", "Date Completed",
    "Close-Out File", "Notes",
]
for col, h in enumerate(headers, 1):
    c = cat.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center

example_rows = [
    [1, "SK-XX-EXAMPLE-1",
     "Wire Instinct + Risk Tolerance into submission finish-bonus",
     "Phase 1", "Stage 1b.1", "M", "SK-XX (✓)",
     "Queued", "Full Spec",
     "Wires the Instinct and Risk Tolerance attributes into the submission finish-bonus calculation.",
     "sk-xx-wave-a", "", "", "",
     "Example row — replace with real sprints during /roadmap-review."],
    [2, "SK-XX-EXAMPLE-2",
     "CLAUDE.md archive pass + branch cleanup",
     "Phase 1", "Stage 1b.2", "S", "",
     "Queued", "Full Spec",
     "Trims project context document below soft size limit; merges architectural memo to main.",
     "doc-arch-cleanup", "", "", "", "Example row."],
    [3, "SK-XX-EXAMPLE-3",
     "Experience into willpower decay (example stub)",
     "Phase 1", "Stage 1c", "M", "SK-XX-EXAMPLE-2",
     "Queued", "Stub",
     "Stub for future sprint — promoted to Full Spec at next /roadmap-review.",
     "", "", "", "", "Example row."],
    ["", "SK-XX-EXAMPLE-DONE",
     "Past sprint that has already shipped",
     "Phase 1", "Stage 1a", "M", "",
     "Completed", "Full Spec",
     "Example of a completed sprint — note Seq is blank.",
     "sk-xx-done-wave-a", "2026-04-20", "2026-04-22",
     "CloseOut.SK-XX-EXAMPLE-DONE.2026-04-22.md", "Example row."],
]
for row in example_rows:
    cat.append(row)

impl_dv = DataValidation(
    type="list",
    formula1='"Queued,In Flight,Blocked,Completed,Dissolved"',
    allow_blank=True, showDropDown=False)
impl_dv.error = "Pick a valid Implementation Status"
impl_dv.prompt = "Queued / In Flight / Blocked / Completed / Dissolved"
cat.add_data_validation(impl_dv)
impl_dv.add("H2:H2000")

wstat_dv = DataValidation(
    type="list",
    formula1='"None,Stub,Draft,Skeleton,Full Spec"',
    allow_blank=True, showDropDown=False)
wstat_dv.error = "Pick a valid Written Status"
wstat_dv.prompt = "None / Stub / Draft / Skeleton / Full Spec"
cat.add_data_validation(wstat_dv)
wstat_dv.add("I2:I2000")

loe_dv = DataValidation(
    type="list",
    formula1='"S,M,L,XL"',
    allow_blank=True, showDropDown=False)
loe_dv.error = "Pick S, M, L, or XL"
loe_dv.prompt = "T-shirt sizing"
cat.add_data_validation(loe_dv)
loe_dv.add("F2:F2000")

last_row = 1 + len(example_rows)
last_col_letter = get_column_letter(len(headers))
tbl = Table(displayName="SprintCatalog", ref=f"A1:{last_col_letter}{last_row}")
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False)
cat.add_table(tbl)

widths = [6, 18, 50, 12, 14, 6, 18, 22, 16, 60, 22, 14, 14, 36, 30]
for col, w in enumerate(widths, 1):
    cat.column_dimensions[get_column_letter(col)].width = w

for row in cat.iter_rows(min_row=2, max_row=last_row, min_col=10, max_col=10):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

cat.freeze_panes = "A2"

wb.save(OUT)
print(f"Wrote {OUT}")
