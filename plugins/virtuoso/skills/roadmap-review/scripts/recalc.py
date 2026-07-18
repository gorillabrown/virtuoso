"""Sync sprint-catalog.csv row data into sprint-queue.xlsx's Catalog tab.

sprint-catalog.csv is the single source of truth for sprint data (see the
roadmap-review, roadmap-status, and next-pointer skills). sprint-queue.xlsx is an
optional, human-facing report — this script keeps its Catalog tab's data cells in
sync with the CSV so the workbook opens with current sprint data. It never writes
Dashboard cells: the Dashboard is a Power Query / formula view over Catalog that
Excel recomputes on its own when a human opens the workbook, and force-writing
those cells via openpyxl (the prior behavior of this script) fights that refresh
and produces a cache that silently disagrees with both Excel and the CSV.

Usage:
    python recalc.py <path-to-sprint-catalog.csv> <path-to-sprint-queue.xlsx>
"""
import csv
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CATALOG_HEADERS = [
    "Seq", "Sprint Code", "Phase", "Stage", "Title", "LOE", "Dependencies",
    "Implementation Status", "Written Status", "Branch", "Date Started",
    "Date Completed", "Close-Out File", "Description", "Notes",
]


def _read_csv_rows(csv_path):
    with open(csv_path, newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return [row for row in reader if (row.get("Sprint Code") or "").strip()]


def _find_catalog(wb):
    """Locate the sprint catalog sheet by name, else by a 'Sprint Code' header
    (robust to the data sheet being named Catalog, DATA.sprint-catalog, etc.)."""
    for name in wb.sheetnames:
        if name.strip().lower() in ("catalog", "data.sprint-catalog", "sprint catalog"):
            return wb[name]
    for ws in wb.worksheets:
        if any(isinstance(c.value, str) and c.value.strip() == "Sprint Code" for c in ws[1]):
            return ws
    raise KeyError("No catalog sheet (need a 'Sprint Code' header in row 1)")


def sync_catalog(csv_path, xlsx_path):
    rows = _read_csv_rows(csv_path)
    wb = load_workbook(xlsx_path)
    cat = _find_catalog(wb)

    headers = [c.value.strip() if isinstance(c.value, str) else c.value for c in cat[1]]
    col_by_header = {h: i + 1 for i, h in enumerate(headers) if h}

    # Clear existing data rows below the header, then rewrite from the CSV.
    if cat.max_row > 1:
        cat.delete_rows(2, cat.max_row - 1)
    for r, row in enumerate(rows, start=2):
        for header, col in col_by_header.items():
            cat.cell(row=r, column=col, value=row.get(header, ""))

    # Keep the Catalog table's range in sync so Excel formulas/Power Query see all rows.
    last_row = max(len(rows) + 1, 1)
    last_col_letter = get_column_letter(len(headers))
    for table in cat.tables.values():
        first_cell = table.ref.split(":")[0]
        table.ref = f"{first_cell}:{last_col_letter}{last_row}"

    wb.save(xlsx_path)
    return len(rows)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(
            "usage: python recalc.py <path-to-sprint-catalog.csv> <path-to-sprint-queue.xlsx>",
            file=sys.stderr,
        )
        sys.exit(2)
    n = sync_catalog(sys.argv[1], sys.argv[2])
    print(f"synced {n} catalog rows from {sys.argv[1]} into {sys.argv[2]}")
