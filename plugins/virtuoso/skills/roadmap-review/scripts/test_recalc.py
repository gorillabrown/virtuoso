import csv
import os
import subprocess
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = os.path.dirname(os.path.abspath(__file__))

CATALOG_HEADERS = [
    "Seq", "Sprint Code", "Phase", "Stage", "Title", "LOE", "Dependencies",
    "Implementation Status", "Written Status", "Branch", "Date Started",
    "Date Completed", "Close-Out File", "Description", "Notes",
]

DASHBOARD_FORMULA = '=COUNTA(Catalog!B:B)-1'


def _make_workbook(path, example_rows):
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    dash["B11"] = DASHBOARD_FORMULA  # a live formula; must survive sync untouched

    cat = wb.create_sheet("Catalog")
    cat.append(CATALOG_HEADERS)
    for row in example_rows:
        cat.append(row)
    tbl = Table(displayName="SprintCatalog", ref=f"A1:O{1 + len(example_rows)}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    cat.add_table(tbl)

    wb.save(path)


def _write_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(CATALOG_HEADERS)
        for row in rows:
            writer.writerow(row)


def test_sync_catalog_writes_csv_rows_into_catalog_tab(tmp_path):
    from recalc import sync_catalog

    xlsx_path = str(tmp_path / "sprint-queue.xlsx")
    csv_path = str(tmp_path / "sprint-catalog.csv")

    _make_workbook(
        xlsx_path,
        [[1, "EX-1", "Phase 1", "Stage 1", "Example row", "M", "", "Queued", "Stub", "", "", "", "", "", ""]],
    )
    _write_csv(
        csv_path,
        [
            [1, "SK-001", "Phase 1", "Stage 1", "Build parser", "M", "", "Queued", "Full Spec", "sk-001", "", "", "", "Desc", "Note"],
            [2, "SK-002", "Phase 1", "Stage 2", "Stub follow-up", "S", "SK-001", "Blocked", "Stub", "", "", "", "", "", ""],
        ],
    )

    count = sync_catalog(csv_path, xlsx_path)

    assert count == 2
    cat = load_workbook(xlsx_path)["Catalog"]
    rows = list(cat.iter_rows(min_row=2, max_row=cat.max_row, values_only=True))
    assert rows[0][:2] == ("1", "SK-001")
    assert rows[1][:2] == ("2", "SK-002")
    assert rows[0][7] == "Queued"
    assert rows[1][6] == "SK-001"  # Dependencies column
    assert cat.max_row == 3  # header + 2 rows, stale example row gone


def test_sync_catalog_never_touches_dashboard_cells(tmp_path):
    from recalc import sync_catalog

    xlsx_path = str(tmp_path / "sprint-queue.xlsx")
    csv_path = str(tmp_path / "sprint-catalog.csv")

    _make_workbook(xlsx_path, [])
    _write_csv(csv_path, [[1, "SK-001", "Phase 1", "Stage 1", "T", "M", "", "Queued", "Full Spec", "", "", "", "", "", ""]])

    sync_catalog(csv_path, xlsx_path)

    dash = load_workbook(xlsx_path)["Dashboard"]
    assert dash["B11"].value == DASHBOARD_FORMULA


def test_sync_catalog_updates_table_range_to_match_row_count(tmp_path):
    from recalc import sync_catalog

    xlsx_path = str(tmp_path / "sprint-queue.xlsx")
    csv_path = str(tmp_path / "sprint-catalog.csv")

    _make_workbook(
        xlsx_path,
        [[1, "EX-1", "Phase 1", "Stage 1", "Example row", "M", "", "Queued", "Stub", "", "", "", "", "", ""]],
    )
    _write_csv(
        csv_path,
        [
            [1, "SK-001", "Phase 1", "Stage 1", "T", "M", "", "Queued", "Full Spec", "", "", "", "", "", ""],
            [2, "SK-002", "Phase 1", "Stage 1", "T", "S", "", "Queued", "Full Spec", "", "", "", "", "", ""],
            [3, "SK-003", "Phase 1", "Stage 1", "T", "S", "", "Queued", "Full Spec", "", "", "", "", "", ""],
        ],
    )

    sync_catalog(csv_path, xlsx_path)

    cat = load_workbook(xlsx_path)["Catalog"]
    table = cat.tables["SprintCatalog"]
    assert table.ref == "A1:O4"


def test_sync_catalog_skips_rows_without_sprint_code(tmp_path):
    from recalc import sync_catalog

    xlsx_path = str(tmp_path / "sprint-queue.xlsx")
    csv_path = str(tmp_path / "sprint-catalog.csv")

    _make_workbook(xlsx_path, [])
    _write_csv(
        csv_path,
        [
            ["", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            [1, "SK-001", "Phase 1", "Stage 1", "T", "M", "", "Queued", "Full Spec", "", "", "", "", "", ""],
        ],
    )

    count = sync_catalog(csv_path, xlsx_path)

    assert count == 1


def test_cli_invocation_syncs_catalog(tmp_path):
    xlsx_path = str(tmp_path / "sprint-queue.xlsx")
    csv_path = str(tmp_path / "sprint-catalog.csv")

    _make_workbook(xlsx_path, [])
    _write_csv(csv_path, [[1, "SK-001", "Phase 1", "Stage 1", "T", "M", "", "Queued", "Full Spec", "", "", "", "", "", ""]])

    result = subprocess.run(
        [sys.executable, os.path.join(HERE, "recalc.py"), csv_path, xlsx_path],
        check=True, capture_output=True, text=True,
    )

    assert "synced 1 catalog rows" in result.stdout
    cat = load_workbook(xlsx_path)["Catalog"]
    assert cat["B2"].value == "SK-001"
