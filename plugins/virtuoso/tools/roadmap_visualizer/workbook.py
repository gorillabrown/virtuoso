from __future__ import annotations

import datetime as dt
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .model import SprintRow


DASHBOARD_CELLS = {
    "total": "B11",
    "completed": "B12",
    "blocked": "B13",
    "queued": "B14",
    "in_flight": "B15",
    "dissolved": "B16",
    "superseded": "B17",
    "percent_count_complete": "B18",
    "remaining_points": "B21",
    "completed_points": "B22",
    "total_points": "B23",
    "percent_points_complete": "B24",
    "sprints_remaining": "B25",
    "average_points": "B26",
}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).strip()


def _as_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _clean(value).lower() in {"true", "yes", "y", "1", "done"}


def _dependencies(value: Any) -> list[str]:
    text = _clean(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[,;\n]+", text) if part.strip()]


def _find_catalog(wb: Any) -> Any:
    catalog_names = {"data.sprint-catalog", "catalog", "sprint catalog"}
    for name in wb.sheetnames:
        if name.strip().lower() in catalog_names:
            return wb[name]

    for ws in wb.worksheets:
        headers = [_clean(cell.value) for cell in ws[1]]
        if "Sprint Code" in headers:
            return ws

    raise KeyError("No sprint catalog sheet with a 'Sprint Code' header")


def _row_value(row: tuple[Any, ...], index: dict[str, int], name: str) -> Any:
    column = index.get(name)
    if column is None or column >= len(row):
        return None
    return row[column]


def read_workbook(path: Path | str) -> tuple[list[SprintRow], dict[str, Any]]:
    wb = load_workbook(Path(path), read_only=True, data_only=True)
    catalog = _find_catalog(wb)
    headers = [_clean(cell.value) for cell in catalog[1]]
    index = {name: i for i, name in enumerate(headers) if name}

    rows: list[SprintRow] = []
    for values in catalog.iter_rows(min_row=2, values_only=True):
        code = _clean(_row_value(values, index, "Sprint Code"))
        if not code:
            continue
        rows.append(
            SprintRow(
                priority=_clean(_row_value(values, index, "Priority")),
                seq=_as_int(_row_value(values, index, "Seq")),
                code=code,
                phase=_clean(_row_value(values, index, "Phase")),
                stage=_clean(_row_value(values, index, "Stage")),
                title=_clean(_row_value(values, index, "Title")),
                loe=_clean(_row_value(values, index, "LOE")),
                dependencies=_dependencies(_row_value(values, index, "Dependencies")),
                implementation_status=_clean(
                    _row_value(values, index, "Implementation Status")
                ),
                written_status=_clean(_row_value(values, index, "Written Status")),
                branch=_clean(_row_value(values, index, "Branch")),
                date_started=_clean(_row_value(values, index, "Date Started")),
                date_completed=_clean(_row_value(values, index, "Date Completed")),
                close_out_file=_clean(_row_value(values, index, "Close-Out File")),
                description=_clean(_row_value(values, index, "Description")),
                notes=_clean(_row_value(values, index, "Notes")),
                done=_as_bool(_row_value(values, index, "Done?")),
                sort_key=_clean(_row_value(values, index, "SortKey")),
            )
        )

    dashboard: dict[str, Any] = {}
    if "Dashboard" in wb.sheetnames:
        dashboard_ws = wb["Dashboard"]
        for name, cell in DASHBOARD_CELLS.items():
            dashboard[name] = dashboard_ws[cell].value

    return rows, dashboard
