"""Spreadsheet work register (item 23), with a declared runtime dependency (item 79).

``openpyxl`` is optional. When it is absent the provider still constructs — so a
registry that names it validates — but every capability is withdrawn and the
negotiation error names the missing dependency instead of raising ImportError
somewhere deep in a ceremony.
"""
from __future__ import annotations

import os

from .. import textio
from . import base, mapping as mapping_mod

DEPENDENCY = "openpyxl"


def dependency_available() -> tuple[bool, str]:
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        return False, str(exc)
    return True, ""


class XlsxWorkRegister(base.WorkRegisterProvider):
    name = "xlsx"

    def __init__(self, *, source: str, mapping=None, read_only: bool = False,
                 sheet: str = "") -> None:
        super().__init__(source=source, mapping=mapping or mapping_mod.Mapping(),
                         read_only=read_only)
        self.sheet = sheet
        self._available, self._reason = dependency_available()

    @property
    def capabilities(self) -> frozenset[str]:
        if not self._available:
            return frozenset()
        reads = {base.LIST_ACTIVE, base.READ_SEQUENCE, base.READ_STATUS,
                 base.READ_PREREQUISITES, base.READ_EFFORT, base.NEXT_ELIGIBLE}
        return frozenset(reads | {base.WRITE_STATUS, base.STORE_SPEC_LINK,
                                  base.RECORD_COMPLETION})

    def require(self, *capabilities: str) -> None:
        if not self._available:
            raise base.CapabilityError(
                "the spreadsheet provider needs %s, which is not installed (%s). "
                "Install it or register a different work-register provider."
                % (DEPENDENCY, self._reason),
                detail={"dependency": DEPENDENCY})
        super().require(*capabilities)

    def describe(self) -> dict:
        data = super().describe()
        data["dependency"] = {"name": DEPENDENCY, "available": self._available}
        if not self._available:
            data["dependency"]["reason"] = self._reason
        return data

    # -- io ------------------------------------------------------------------

    def _load(self, *, data_only: bool = True):
        from openpyxl import load_workbook
        return load_workbook(self.source, data_only=data_only)

    def _worksheet(self, workbook):
        if self.sheet and self.sheet in workbook.sheetnames:
            return workbook[self.sheet]
        for worksheet in workbook.worksheets:
            headers = [str(cell.value or "").strip() for cell in worksheet[1]]
            if "id" in self.mapping.fields.resolve_index(headers):
                return worksheet
        raise KeyError("no sheet in %s carries an identifiable id column" % self.source)

    def snapshot(self) -> base.Snapshot:
        self.require(base.LIST_ACTIVE)
        workbook = self._load()
        worksheet = self._worksheet(workbook)
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        index = self.mapping.fields.resolve_index(headers)
        items = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            item = self._to_item(row, index)
            if item.id:
                items.append(item)
        return base.Snapshot(items=items, provider=self.name, source=self.source,
                             taken_at=base.utc_now(), fields=sorted(index))

    def _to_item(self, row, index) -> base.WorkItem:
        def value(name: str) -> str:
            position = index.get(name)
            if position is None or position >= len(row):
                return ""
            cell = row[position]
            if cell is None:
                return ""
            if hasattr(cell, "isoformat"):
                return cell.isoformat()[:10]
            return str(cell).strip()

        sequence_raw = value("sequence")
        try:
            sequence = int(float(sequence_raw)) if sequence_raw else None
        except ValueError:
            sequence = None
        raw_status = value("status")
        return base.WorkItem(
            id=value("id"), title=value("title"), sequence=sequence,
            status=self.mapping.statuses.to_canonical(raw_status), raw_status=raw_status,
            written_status=self.mapping.statuses.written_to_canonical(value("written_status")),
            raw_written_status=value("written_status"),
            prerequisites=[p.strip() for p in
                           value("prerequisites").replace(";", ",").split(",") if p.strip()],
            effort=value("effort"), lane=value("lane"), group=value("group"),
            spec_link=value("spec_link"), branch=value("branch"), started=value("started"),
            completed=value("completed"), evidence=value("evidence"),
            description=value("description"), notes=value("notes"),
            revision=textio.sha256_bytes(
                "|".join(str(c) for c in row).encode("utf-8")),
        )

    # -- mutations -----------------------------------------------------------

    def _write_field(self, item_id: str, field_name: str, value: str,
                     revision: str) -> base.WorkItem:
        if self.read_only:
            raise base.CapabilityError("work register %s is registered read-only" % self.source)
        workbook = self._load(data_only=False)
        worksheet = self._worksheet(workbook)
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        index = self.mapping.fields.resolve_index(headers)
        if field_name not in index:
            raise base.CapabilityError(
                "%s has no %r column; configure policy.workRegister.fieldMappings.%s"
                % (self.source, field_name, field_name))
        id_column = index["id"]
        for row_number in range(2, worksheet.max_row + 1):
            values = [worksheet.cell(row=row_number, column=c + 1).value
                      for c in range(len(headers))]
            if str(values[id_column] or "").strip() != item_id:
                continue
            current = self._to_item(tuple(values), index)
            self.check_revision(current, revision)
            cell = worksheet.cell(row=row_number, column=index[field_name] + 1)
            if str(cell.value or "").strip() == value:
                return current            # idempotent
            cell.value = value
            workbook.save(self.source)
            return self.get(item_id)
        raise KeyError("item %r is not in %s" % (item_id, self.source))

    def set_status(self, item_id: str, status: str, *, revision: str = "",
                   raw: str = "") -> base.WorkItem:
        self.require(base.WRITE_STATUS)
        return self._write_field(item_id, "status",
                                 raw or self.mapping.statuses.to_project(status), revision)

    def store_spec_link(self, item_id: str, link: str, *, revision: str = "") -> base.WorkItem:
        self.require(base.STORE_SPEC_LINK)
        return self._write_field(item_id, "spec_link", link, revision)

    def record_completion(self, item_id: str, *, completed: str = "", evidence: str = "",
                          revision: str = "") -> base.WorkItem:
        self.require(base.RECORD_COMPLETION)
        item = self._write_field(item_id, "status",
                                 self.mapping.statuses.to_project(base.COMPLETED), revision)
        if completed:
            item = self._write_field(item_id, "completed", completed, "")
        if evidence:
            item = self._write_field(item_id, "evidence", evidence, "")
        return item

    def exists(self) -> bool:
        return os.path.isfile(self.source)
